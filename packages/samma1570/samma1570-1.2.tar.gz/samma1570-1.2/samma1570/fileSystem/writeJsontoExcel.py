#   @author: 马朝威 1570858572@qq.com
#   @time: 2019-06-02 12:29

import os
import openpyxl
from openpyxl.utils.exceptions import IllegalCharacterError


def write_json_excel(folder_path, file_name):

    if not os.path.exists(folder_path):
        print('folder {} not exists . '.format(folder_path))
        os.mkdir(folder_path)
        print('folder {} has created ! '.format(folder_path))
    else:
        pass

    work_book = openpyxl.Workbook()
    work_sheet = work_book.active

    title_is_none = True
    row_num = 1
    while True:

        json_item = yield
        if json_item is None:
            break

        # 写入字段名
        if title_is_none:
            for i, key in enumerate(json_item.keys()):
                work_sheet.cell(row=row_num, column=i+1, value=key)
            title_is_none = False
            row_num = row_num + 1
        else:
            pass
        # 写入字段
        for j, value in enumerate(json_item.values()):
            try:
                work_sheet.cell(row=row_num, column=j+1, value=value)
            except IllegalCharacterError as IErr:
                print('meet IllegalCharacterError, but pass.')
        row_num = row_num + 1

    file_path = ''.join([folder_path, '\\',file_name, '.xlsx'])
    work_book.save(file_path)
    print('file {} has created ! '.format(file_path))



def main():
    pass


if __name__ == "__main__":
    main()










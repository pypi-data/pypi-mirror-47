import openpyxl
from openpyxl.cell import MergedCell
from openpyxl.utils import get_column_letter # int_to_column_letter
from openpyxl.utils import column_index_from_string # column_letter_to_int
from openpyxl.utils import range_boundaries # (A1:B2) ==> (1, 1, 2, 2)


COLUMN_LETTER_A_VALUE = ord('A')
COLUMN_LETTER_BASE_VALUE = COLUMN_LETTER_A_VALUE - 1 
MAX_ROW_INDEX = 65536   # 这个值EXCEL本身并无限制
MAX_COL_INDEX = 18278   # column_index_from_string('ZZZ65536')，这个值EXCEL是有限制的


def get_workbook(workbook, **kwargs):
    """获取Workbook对象。
    """
    if isinstance(workbook, str):
        workbook = openpyxl.load_workbook(workbook, **kwargs)
    return workbook


def get_worksheet(workbook, sheet=None):
    """根据表单名获取表单对象。
    """
    workbook = get_workbook(workbook)
    worksheet = None
    if not sheet:
        worksheet = workbook.active
    elif isinstance(sheet, int):
        worksheet = workbook[workbook.get_sheet_names()[sheet]]
    else:
        worksheet = workbook[sheet]
    return worksheet


def get_merged_range_string(min_col, min_row, max_col, max_row):
    """合并单元格，元组范围转字符串范围。

    get_merged_range_string(1, 1, 2, 2) == "A1:B2"
    """
    return "{0}{1}:{2}{3}".format(
        get_column_letter(min_col),
        min_row,
        get_column_letter(max_col),
        max_row
    )


def parse_cols(cols, max_col=None):
    """列名列表（支持减号表范围，逗号表并列）解析。

    parse_cols("Z") == [26]
    parse_cols("A,C") == [1, 3]
    parse_cols("A-C") == [1, 2, 3]
    parse_cols("A-C,E,G,I-J") == [1, 2, 3, 5, 7, 9, 10]
    """
    if not cols:
        return []
    if isinstance(cols, (list, tuple)):
        return list(cols)
    ns = []
    ps = [x.strip() for x in cols.split(",")]
    for p in ps:
        if not "-" in p:
            ns.append(column_index_from_string(p))
        else:
            s, e = [x.strip() for x in p.split("-")]
            sv = column_index_from_string(s)
            if not e:
                if max_col:
                    ev = max_col
                else:
                    ev = MAX_COL_INDEX
            else:
                ev = column_index_from_string(e)
            ns += list(range(sv, ev+1))
    return ns

def parse_rows(rows, max_row=None):
    """行列表解析。

    parse_rows("1") == [1]
    parse_rows("1-3") == [1, 2, 3]
    parse_rows("1-3,4") == [1, 2, 3, 4]
    parse_rows("1-3,4-5") == [1, 2, 3, 4, 5]
    parse_rows("1-3,4-5,8,10,13") == [1, 2, 3, 4, 5, 8, 10, 13]
    """
    if not rows:
        return []
    if isinstance(rows, (list, tuple)):
        return list(rows)
    ns = []
    ps = [x.strip() for x in rows.split(",")]
    for p in ps:
        if not "-" in p:
            ns.append(int(p))
        else:
            s, e = [x.strip() for x in p.split("-")]
            sv = int(s)
            if not e:
                if max_row:
                    ev = max_row
                else:
                    ev = MAX_ROW_INDEX
            else:
                ev = int(e)
            ns += list(range(sv, ev+1))
    return ns


def get_cells(worksheet, rows, cols):
    """根据行列，获取子表。
    """
    table = []
    row_values = parse_rows(rows)
    col_values = parse_cols(cols)
    for row_index in row_values:
        row = []
        for col_index in col_values:
            cell = worksheet.cell(row_index, col_index)
            row.append(cell)
        table.append(row)
    return table


def get_merged_ranges(worksheet, rows, cols):
    """计算子表中合并单元格的范围列表。
    """
    if not rows or not cols:
        return []
    result = []
    row_values = parse_rows(rows)
    col_values = parse_cols(cols)
    merged_ranges = worksheet.merged_cell_ranges
    for merged_range in merged_ranges:
        min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))
        min_col_merged = max(col_values) + 1
        max_col_merged = 0
        min_row_merged = max(row_values) + 1
        max_row_merged = 0
        for col in range(min_col, max_col + 1):
            if col in col_values:
                col_merged = col_values.index(col) + 1
                if min_col_merged > col_merged:
                    min_col_merged = col_merged
                if max_col_merged < col_merged:
                    max_col_merged = col_merged
        for row in range(min_row, max_row + 1):
            if row in row_values:
                row_merged = row_values.index(row) + 1
                if min_row_merged > row_merged:
                    min_row_merged = row_merged
                if max_row_merged < row_merged:
                    max_row_merged = row_merged
        if min_col_merged <= max_col_merged and min_row_merged <= max_row_merged:
            if not (min_col_merged == max_col_merged and min_row_merged == max_row_merged):
                mergered_range = get_merged_range_string(min_col_merged, min_row_merged, max_col_merged, max_row_merged)
                result.append(mergered_range)
    return result

def copy_cells(src_cells, dst_worksheet, top=1, left=1):
    row_delta = 0
    for row in src_cells:
        row_delta += 1
        row_idx = top + row_delta - 1
        col_delta = 0
        for src_cell in row:
            col_delta += 1
            col_idx = left + col_delta - 1
            dst_cell = dst_worksheet.cell(row_idx, col_idx, src_cell.value)
            dst_cell.alignment = src_cell.alignment.copy()
            dst_cell.border = src_cell.border.copy()
            dst_cell.fill = src_cell.fill.copy()
            dst_cell.font = src_cell.font.copy()
            #dst_cell.offset = src_cell.offset.copy()
            #dst_cell.number_format = src_cell.number_format.copy()

def merge_ranges(worksheet, merge_ranges, top=1, left=1):
    top -= 1
    left -= 1
    for merge_range in merge_ranges:
        min_col, min_row, max_col, max_row = range_boundaries(merge_range)
        min_col += left
        max_col += left
        min_row += top
        max_row += top
        merge_range = get_merged_range_string(min_col, min_row, max_col, max_row)
        worksheet.merge_cells(merge_range)

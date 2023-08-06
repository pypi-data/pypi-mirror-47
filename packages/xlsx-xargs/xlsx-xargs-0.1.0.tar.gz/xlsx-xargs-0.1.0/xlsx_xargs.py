import re
import copy
import subprocess
import click
from openpyxl.utils import get_column_letter
from xlsxhelper import get_workbook
from xlsxhelper import get_worksheet
from xlsxhelper import parse_rows


import warnings
warnings.filterwarnings("ignore", module="openpyxl.worksheet._reader")


def xargs(workbook, sheet, rows, tests, commands):
    """类似xargs命令，遍历依据是excel表单中的每一行。命令参数，允许被string.format替换，替换内容为指定单元格值。
    """
    workbook = get_workbook(workbook, data_only=True)
    worksheet = get_worksheet(workbook, sheet)
    tests = tests or []
    if rows:
        row_values = parse_rows(rows, worksheet.max_row)
    else:
        row_values = list(range(1, worksheet.max_row + 1))
    rows = list(worksheet.rows)
    for row_index in row_values:
        row = rows[row_index-1]
        col_index = 0
        data = {}
        for cell in row:
            col_index += 1
            col_letter = get_column_letter(col_index)
            data[col_letter] = cell.value
        test_ok = True
        for test in tests:
            col_letter, regex = test.split(":")
            col_value = str(data.get(col_letter, ""))
            if not re.match(regex, col_value):
                test_ok = False
                break
        if not test_ok:
            continue
        cmd_args = copy.copy(list(commands))
        for idx in range(len(cmd_args)):
            cmd_args[idx] = cmd_args[idx].format(**data)
        subprocess.check_call(cmd_args, shell=True)


@click.command()
@click.option("-f", "--file", required=True, help="Excel文件路径。")
@click.option("-s", "--sheet", help="表单页名称。默认为当前页。")
@click.option("-r", "--rows", help="指定需要处理的行。默认为所有行。")
@click.option("-t", "--test", multiple=True, help="使用正则对行进行过滤，只处理匹配的行。允许使用多个表达式，多个表达式之间求与。")
@click.argument("commands", nargs=-1)
def main(file, sheet, rows, test, commands):
    """类似xargs命令，遍历依据是excel表单中的每一行。命令参数，允许被string.format替换，替换内容为指定单元格值。

    注意：

    COMMANDS前加两个减号，可表示减号后均为COMMANDS参数。
    xlsx-xargs [OPTIONS] -- [COMMANDS]...
    """
    xargs(file, sheet, rows, test, commands)


if __name__ == "__main__":
    main()
